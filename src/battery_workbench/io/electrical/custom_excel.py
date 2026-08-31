from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass(frozen=True)
class SheetInfo:
    rows: int
    columns: int


@dataclass(frozen=True)
class ElectricalWorkbookInspection:
    sheets: dict[str, SheetInfo]


@dataclass(frozen=True)
class RawSheetRow:
    source_row_index: int
    values: dict[str, Any]


@dataclass(frozen=True)
class RawSheetData:
    name: str
    info: SheetInfo
    header_row: int
    headers: tuple[str, ...]
    rows: tuple[RawSheetRow, ...]


@dataclass(frozen=True)
class ElectricalWorkbookData:
    sheets: dict[str, RawSheetData]


SHEET_HEADER_ROWS = {
    "record": 1,
    "cycle": 1,
    "step": 1,
    "auxTemp": 2,
    "auxVol": 2,
}


def inspect_electrical_workbook(path: str | Path) -> ElectricalWorkbookInspection:
    """Schema-level inspection only. Does not mutate the source workbook."""
    path = Path(path)
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        return ElectricalWorkbookInspection(
            sheets={
                name: SheetInfo(rows=wb[name].max_row, columns=wb[name].max_column)
                for name in wb.sheetnames
            }
        )
    finally:
        wb.close()


def read_electrical_workbook(path: str | Path) -> ElectricalWorkbookData:
    """Read parser-relevant sheets without modifying the workbook."""
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    try:
        sheets: dict[str, RawSheetData] = {}
        for sheet_name, header_row in SHEET_HEADER_ROWS.items():
            if sheet_name not in workbook.sheetnames:
                continue
            worksheet = workbook[sheet_name]
            header_values = next(
                worksheet.iter_rows(
                    min_row=header_row,
                    max_row=header_row,
                    values_only=True,
                )
            )
            headers = tuple(
                str(value).strip() if value is not None else f"__unnamed_{index}"
                for index, value in enumerate(header_values, start=1)
            )
            rows = tuple(
                RawSheetRow(
                    source_row_index=row_index,
                    values=dict(zip(headers, values, strict=True)),
                )
                for row_index, values in enumerate(
                    worksheet.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                )
            )
            sheets[sheet_name] = RawSheetData(
                name=sheet_name,
                info=SheetInfo(rows=worksheet.max_row, columns=worksheet.max_column),
                header_row=header_row,
                headers=headers,
                rows=rows,
            )
        return ElectricalWorkbookData(sheets=sheets)
    finally:
        workbook.close()
