from __future__ import annotations

import json
from collections.abc import Callable
from datetime import datetime, timedelta
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook
from openpyxl.styles import PatternFill

RECORD_HEADERS = [
    "数据序号",
    "循环号",
    "工步号",
    "工步开始结束标识",
    "工步类型",
    "时间",
    "总时间",
    "电流(A)",
    "电压(V)",
    "容量(Ah)",
    "充电容量(Ah)",
    "放电容量(Ah)",
    "能量(Wh)",
    "充电能量(Wh)",
    "放电能量(Wh)",
    "绝对时间",
    "功率(W)",
    "dQ/dV(mAh/V)",
    "接触电阻(mΩ)",
    "SOC/DOD(%)",
]


def _write_electrical_workbook(
    path: Path,
    *,
    start: datetime,
    missing_record_column: str | None = None,
    backwards_timestamp: bool = False,
    include_aux_voltage: bool = True,
) -> Path:
    workbook = Workbook()
    record = workbook.active
    record.title = "record"
    headers = [header for header in RECORD_HEADERS if header != missing_record_column]
    record.append(headers)

    timestamps = [start, start + timedelta(seconds=1), start + timedelta(seconds=1), start]
    if not backwards_timestamp:
        timestamps[-1] = start + timedelta(seconds=2)

    rows = [
        {
            "数据序号": 1,
            "循环号": 1,
            "工步号": 1,
            "工步开始结束标识": 0,
            "工步类型": "恒流充电",
            "时间": "00:00:00",
            "总时间": "00:00:00",
            "电流(A)": 1.0,
            "电压(V)": 3.2,
            "容量(Ah)": 0.0,
            "充电容量(Ah)": 0.0,
            "放电容量(Ah)": 0.0,
            "能量(Wh)": 0.0,
            "充电能量(Wh)": 0.0,
            "放电能量(Wh)": 0.0,
            "绝对时间": timestamps[0].strftime("%Y-%m-%d %H:%M:%S"),
            "功率(W)": 3.2,
            "dQ/dV(mAh/V)": 0.0,
            "接触电阻(mΩ)": 1.5,
            "SOC/DOD(%)": 0.0,
        },
        {
            "数据序号": 2,
            "循环号": 1,
            "工步号": 2,
            "工步开始结束标识": 0,
            "工步类型": "搁置",
            "时间": "00:00:01",
            "总时间": "00:00:01",
            "电流(A)": 0.0,
            "电压(V)": 3.3,
            "容量(Ah)": 0.1,
            "充电容量(Ah)": 0.1,
            "放电容量(Ah)": 0.0,
            "能量(Wh)": 0.33,
            "充电能量(Wh)": 0.33,
            "放电能量(Wh)": 0.0,
            "绝对时间": timestamps[1].strftime("%Y-%m-%d %H:%M:%S"),
            "功率(W)": 0.0,
            "dQ/dV(mAh/V)": 1.0,
            "接触电阻(mΩ)": 1.4,
            "SOC/DOD(%)": 10.0,
        },
        {
            "数据序号": 3,
            "循环号": 2,
            "工步号": 1,
            "工步开始结束标识": 0,
            "工步类型": "恒流充电",
            "时间": "00:00:00",
            "总时间": "00:00:01",
            "电流(A)": 1.0,
            "电压(V)": 3.4,
            "容量(Ah)": 0.0,
            "充电容量(Ah)": 0.0,
            "放电容量(Ah)": 0.0,
            "能量(Wh)": 0.0,
            "充电能量(Wh)": 0.0,
            "放电能量(Wh)": 0.0,
            "绝对时间": timestamps[2].strftime("%Y-%m-%d %H:%M:%S"),
            "功率(W)": 3.4,
            "dQ/dV(mAh/V)": 0.0,
            "接触电阻(mΩ)": 1.3,
            "SOC/DOD(%)": 0.0,
        },
        {
            "数据序号": 4,
            "循环号": 2,
            "工步号": 2,
            "工步开始结束标识": 0,
            "工步类型": "搁置",
            "时间": "00:00:01",
            "总时间": "00:00:02",
            "电流(A)": 0.0,
            "电压(V)": 3.5,
            "容量(Ah)": 0.1,
            "充电容量(Ah)": 0.1,
            "放电容量(Ah)": 0.0,
            "能量(Wh)": 0.35,
            "充电能量(Wh)": 0.35,
            "放电能量(Wh)": 0.0,
            "绝对时间": timestamps[3].strftime("%Y-%m-%d %H:%M:%S"),
            "功率(W)": 0.0,
            "dQ/dV(mAh/V)": 1.0,
            "接触电阻(mΩ)": 1.2,
            "SOC/DOD(%)": 10.0,
        },
    ]
    for values in rows:
        record.append([values[header] for header in headers])

    cycle = workbook.create_sheet("cycle")
    cycle.append(["循环号", "起始绝对时间", "结束绝对时间", "充电容量(Ah)", "放电容量(Ah)"])
    cycle.append([1, rows[0]["绝对时间"], rows[1]["绝对时间"], 0.1, 0.0])
    cycle.append([2, rows[2]["绝对时间"], rows[3]["绝对时间"], 0.1, 0.0])

    step = workbook.create_sheet("step")
    step.append(
        [
            "循环号",
            "工步号",
            "工步序号",
            "工步类型",
            "工步时间",
            "起始绝对时间",
            "结束绝对时间",
            "容量(Ah)",
            "起始电流(A)",
            "结束电流(A)",
            "DCIR(mΩ)",
        ]
    )
    for index, values in enumerate(rows, start=1):
        step.append(
            [
                values["循环号"],
                values["工步号"],
                index,
                values["工步类型"],
                values["时间"],
                values["绝对时间"],
                values["绝对时间"],
                values["容量(Ah)"],
                values["电流(A)"],
                values["电流(A)"],
                values["接触电阻(mΩ)"],
            ]
        )
    step.cell(row=6, column=1).fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    step.cell(row=7, column=8, value=99.0)

    aux_temp = workbook.create_sheet("auxTemp")
    aux_temp.append([None, None, "单体温度(℃)", None])
    aux_temp.append(["数据序号", "绝对时间", "T1", "辅助通道温差"])
    for index, timestamp in enumerate(timestamps, start=1):
        aux_temp.append([index, timestamp.strftime("%Y-%m-%d %H:%M:%S"), 25.0 + index, 0.0])

    if include_aux_voltage:
        aux_voltage = workbook.create_sheet("auxVol")
        aux_voltage.append([None, None, "单体电压(V)", None])
        aux_voltage.append(["数据序号", "绝对时间", "V1", "辅助通道电压差"])
        for index, timestamp in enumerate(timestamps, start=1):
            aux_voltage.append(
                [index, timestamp.strftime("%Y-%m-%d %H:%M:%S"), 3.0 + index / 10, 0.0]
            )

    workbook.save(path)
    return path


@pytest.fixture
def electrical_workbook_factory(tmp_path: Path) -> Callable[..., Path]:
    def factory(
        name: str = "sample.xlsx",
        *,
        start: datetime | None = None,
        missing_record_column: str | None = None,
        backwards_timestamp: bool = False,
        include_aux_voltage: bool = True,
    ) -> Path:
        return _write_electrical_workbook(
            tmp_path / name,
            start=start or datetime.fromisoformat("2024-01-01 10:00:00"),
            missing_record_column=missing_record_column,
            backwards_timestamp=backwards_timestamp,
            include_aux_voltage=include_aux_voltage,
        )

    return factory


@pytest.fixture
def electrical_qa_input_factory(tmp_path: Path) -> Callable[..., Path]:
    """Write a minimal BRW-003 output bundle without invoking the XLSX parser."""

    def factory(
        *,
        duplicate: bool = False,
        backwards: bool = False,
        large_gap: bool = False,
        missing_column: str | None = None,
        missing_aux_temperature: bool = False,
        cycle_mismatch: bool = False,
        step_mismatch: bool = False,
        voltage_outlier: bool = False,
    ) -> Path:
        input_dir = tmp_path / f"input-{len(list(tmp_path.iterdir()))}"
        input_dir.mkdir()
        start = pd.Timestamp("2024-01-01 00:00:00")
        seconds = [0, 1, 2, 3]
        if duplicate:
            seconds[2] = 1
        if backwards:
            seconds[2] = 0
        if large_gap:
            seconds[2:] = [31, 32]
        timestamps = [start + pd.Timedelta(seconds=value) for value in seconds]
        cycles_raw = [1, 1, 2, 2]
        if cycle_mismatch:
            cycles_raw[-1] = 3
        steps_raw = [1, 2, 1, 2]
        if step_mismatch:
            steps_raw[-1] = 9
        records = pd.DataFrame(
            {
                "battery_id": ["CELL_TEST"] * 4,
                "experiment_id": ["EXP_TEST"] * 4,
                "electrical_asset_id": ["E_TEST"] * 4,
                "source_file": ["test.xlsx"] * 4,
                "source_sheet": ["record"] * 4,
                "source_row_index": [2, 3, 4, 5],
                "record_index_raw": [1, 2, 3, 4],
                "cycle_index_raw": cycles_raw,
                "step_index_raw": steps_raw,
                "step_type_raw": ["CC", "REST", "CC", "REST"],
                "timestamp": timestamps,
                "current_a": [1.0, 0.0, 1.0, 0.0],
                "voltage_v": [99.0 if voltage_outlier else 3.0, 3.1, 3.2, 3.3],
                "capacity_ah": [0.0, 0.1, 0.0, 0.1],
                "charge_capacity_ah": [0.0, 0.1, 0.0, 0.1],
                "discharge_capacity_ah": [0.0] * 4,
                "dqdv_mah_per_v": [0.0, 1.0, 0.0, 1.0],
                "contact_resistance_mohm": [1.0] * 4,
                "soc_dod_percent": [0.0, 10.0, 0.0, 10.0],
            }
        )
        if missing_column:
            records = records.drop(columns=[missing_column])
        cycles = pd.DataFrame(
            {
                "battery_id": ["CELL_TEST"] * 2,
                "experiment_id": ["EXP_TEST"] * 2,
                "electrical_asset_id": ["E_TEST"] * 2,
                "source_file": ["test.xlsx"] * 2,
                "source_sheet": ["cycle"] * 2,
                "source_row_index": [2, 3],
                "cycle_index_raw": [1, 2],
                "start_timestamp": [timestamps[0], timestamps[2]],
                "end_timestamp": [timestamps[1], timestamps[3]],
                "charge_capacity_ah": [0.1, 0.1],
                "discharge_capacity_ah": [0.0, 0.0],
            }
        )
        steps = pd.DataFrame(
            {
                "battery_id": ["CELL_TEST"] * 4,
                "experiment_id": ["EXP_TEST"] * 4,
                "electrical_asset_id": ["E_TEST"] * 4,
                "source_file": ["test.xlsx"] * 4,
                "source_sheet": ["step"] * 4,
                "source_row_index": [2, 3, 4, 5],
                "cycle_index_raw": [1, 1, 2, 2],
                "step_index_raw": [1, 2, 1, 2],
                "step_sequence_raw": [1, 2, 3, 4],
                "step_type_raw": ["CC", "REST", "CC", "REST"],
                "start_timestamp": timestamps,
                "end_timestamp": timestamps,
                "step_time_s": [0.0] * 4,
            }
        )
        aux_common = {
            "battery_id": ["CELL_TEST"] * 4,
            "experiment_id": ["EXP_TEST"] * 4,
            "electrical_asset_id": ["E_TEST"] * 4,
            "source_file": ["test.xlsx"] * 4,
            "source_row_index": [2, 3, 4, 5],
            "record_index_raw": [1, 2, 3, 4],
            "timestamp": timestamps,
        }
        if not missing_aux_temperature:
            pd.DataFrame(
                aux_common
                | {
                    "source_sheet": ["auxTemp"] * 4,
                    "temperature_channel": ["T1"] * 4,
                    "temperature_c": [25.0, 25.1, 25.2, 25.3],
                }
            ).to_parquet(input_dir / "aux_temperature.parquet", index=False)
        pd.DataFrame(
            aux_common
            | {
                "source_sheet": ["auxVol"] * 4,
                "voltage_channel": ["V1"] * 4,
                "voltage_v": [3.0, 3.1, 3.2, 3.3],
            }
        ).to_parquet(input_dir / "aux_voltage.parquet", index=False)
        records.to_parquet(input_dir / "records.parquet", index=False)
        cycles.to_parquet(input_dir / "cycles.parquet", index=False)
        steps.to_parquet(input_dir / "steps.parquet", index=False)
        (input_dir / "parser_manifest.json").write_text(
            json.dumps(
                {
                    "battery_id": "CELL_TEST",
                    "experiment_id": "EXP_TEST",
                    "parser": "custom_excel",
                    "parser_version": "0.1.0",
                    "source_assets": ["E_TEST"],
                }
            ),
            encoding="utf-8",
        )
        return input_dir

    return factory


@pytest.fixture
def ultrasound_txt_factory(tmp_path: Path) -> Callable[..., Path]:
    def factory(
        name: str = "ultrasound.txt",
        *,
        frame_ids: list[int] | None = None,
        elapsed_times: list[float] | None = None,
    ) -> Path:
        ids = frame_ids or [0, 1, 2]
        elapsed = elapsed_times or [0.031, 10.031, 20.031]
        lines = []
        for frame_id, elapsed_time in zip(ids, elapsed, strict=True):
            waveform = " ".join(str(frame_id * 1000 + index) for index in range(1250))
            tail = " ".join(str(index) for index in range(16))
            lines.append(
                f"{frame_id};unknown-{frame_id};{elapsed_time};"
                f"meta-{frame_id} state-{frame_id};{waveform};{tail}\n"
            )
        path = tmp_path / name
        path.write_text("".join(lines), encoding="utf-8")
        return path

    return factory
